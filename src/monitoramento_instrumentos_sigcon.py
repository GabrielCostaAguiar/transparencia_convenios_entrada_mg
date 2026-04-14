import os
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

from config import (
    BASE_DIR, DOWNLOAD_DIR, DIRETORIO_PDFS,
    PLANILHA_CONTROLE, PLANILHA_CONSULTAS_SIGCON,
    PLANILHA_OUTPUT_PREFIX, ABA_EXCEL,
    PLANILHA_LISTA_DOWNLOAD, COL_OBSERVACAO,
    PASTA_DRIVE, PASTA_DRIVE_TA, SCOPES_DRIVE, TOKEN_PATH,
    URL_TRANSFEREGOV_BASE,
    COL_SEI_SIAFI, COL_SEI_INSTRUMENTO,
    COL_SIGCON_SIAFI, COL_SIGCON_CODIGO_UNIAO,
    COL_SIGCON_SITUACAO, COL_SIGCON_DATA_PUB,
    COL_SIGCON_INTEIRO_SIGCON, COL_SIGCON_INTEIRO_TRANSFERE,
    COL_DOC_AUTORIZATIVO, COL_NOME_PDF, COL_ID_PDF, COL_DRIVE_RESOURCE,
    DATA_FILTRO, SITUACAO_BLOQUEADO, COLUNAS_FINAIS, COLUNAS_TA,
    EXCEL_FONTE, EXCEL_TAMANHO, EXCEL_LARGURA_COL, CAMINHO_FINAL
)

load_dotenv(BASE_DIR / ".env")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload


# ── Autenticação ──────────────────────────────────────────────────────────────

def autenticar_drive():
    """Autentica no Google Drive via OAuth2 e retorna o serviço."""
    creds = None

    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES_DRIVE)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                os.getenv("GOOGLE_CREDENTIALS_JSON"),
                SCOPES_DRIVE,
            )
            creds = flow.run_local_server(port=0, login_hint=os.getenv("GOOGLE_AUTH_EMAIL"))

        with open(TOKEN_PATH, "w") as token:
            token.write(creds.to_json())

    print("✅ Autenticado no Google Drive com sucesso!")
    return build("drive", "v3", credentials=creds)


# ── Drive: upload e listagem ──────────────────────────────────────────────────

def _id_pasta_drive(service, pasta_name=PASTA_DRIVE):
    """Retorna o ID da pasta `pasta_name` no Drive."""
    resultado = service.files().list(
        q=f"name='{pasta_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name)",
    ).execute()
    pastas = resultado.get("files", [])
    if not pastas:
        raise RuntimeError(f"Pasta '{pasta_name}' não encontrada no Google Drive.")
    return pastas[0]["id"]


def _nomes_existentes_na_pasta(service, pasta_id):
    resultado = service.files().list(
        q=f"'{pasta_id}' in parents and trashed=false",
        fields="files(name)",
        pageSize=1000,
    ).execute()
    return {f["name"] for f in resultado.get("files", [])}


def _is_termo_aditivo(pdf_path):
    """Retorna True se o arquivo for um Termo Aditivo (stem termina em _<dígitos>)."""
    stem = pdf_path.stem
    partes = stem.rsplit("_", 1)
    return len(partes) == 2 and partes[1].isdigit()


def fazer_upload_pdfs(service):
    """Envia PDFs locais para o Drive: TAs → 'Termos aditivos', demais → 'Instrumento'."""
    pasta_id_inst = _id_pasta_drive(service, PASTA_DRIVE)
    pasta_id_ta   = _id_pasta_drive(service, PASTA_DRIVE_TA)

    nomes_inst = _nomes_existentes_na_pasta(service, pasta_id_inst)
    nomes_ta   = _nomes_existentes_na_pasta(service, pasta_id_ta)

    upados = 0
    for pdf_path in sorted(Path(DIRETORIO_PDFS).glob("*.pdf")):
        if _is_termo_aditivo(pdf_path):
            if pdf_path.name in nomes_ta:
                continue
            pasta_destino = pasta_id_ta
        else:
            if pdf_path.name in nomes_inst:
                continue
            pasta_destino = pasta_id_inst

        file_metadata = {"name": pdf_path.name, "parents": [pasta_destino]}
        media = MediaFileUpload(str(pdf_path), mimetype="application/pdf", resumable=True)
        service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id",
        ).execute()
        print(f"  ↑ Upload: {pdf_path.name}")
        upados += 1

    print(f"✅ {upados} PDF(s) enviado(s) para o Drive.")


def listar_pdfs_drive(service):
    """Lista os arquivos da pasta 'transparencia' e retorna DataFrame com links."""
    pasta_id = _id_pasta_drive(service)

    resultado = service.files().list(
        q=f"'{pasta_id}' in parents and trashed=false",
        fields="files(id, name)",
        pageSize=1000,
    ).execute()
    arquivos = resultado.get("files", [])

    df = pd.DataFrame(arquivos) if arquivos else pd.DataFrame(columns=["id", "name"])
    df[COL_DRIVE_RESOURCE]   = "https://drive.google.com/file/d/" + df["id"] + "/view"
    df[COL_DOC_AUTORIZATIVO] = df["name"].str[:-4]   # remove ".pdf"

    print(f"✅ {len(df)} arquivo(s) listado(s) no Drive (Instrumento).")
    return df


def listar_pdfs_ta_drive(service):
    """Lista os arquivos da pasta 'Termos aditivos' e retorna DataFrame com links."""
    pasta_id = _id_pasta_drive(service, PASTA_DRIVE_TA)

    resultado = service.files().list(
        q=f"'{pasta_id}' in parents and trashed=false",
        fields="files(id, name)",
        pageSize=1000,
    ).execute()
    arquivos = resultado.get("files", [])

    df = pd.DataFrame(arquivos) if arquivos else pd.DataFrame(columns=["id", "name"])
    df["drive_resource_ta"] = "https://drive.google.com/file/d/" + df["id"] + "/view"
    df["stem"]              = df["name"].str[:-4]   # remove ".pdf"

    print(f"✅ {len(df)} arquivo(s) listado(s) no Drive (Termos aditivos).")
    return df


# ── Lista de download ─────────────────────────────────────────────────────────

def gerar_lista_download(sigcon_filtrado, controle_sei):
    """
    Cruza o SIGCON filtrado com o Controle SEI pelo SIAFI e gera um XLSX com:
      - Código SIAFI
      - Doc_autorizativo (= Instrumento do Controle SEI)
      - TA 1 … TA 22 (colunas presentes no Controle SEI)
      - Observação (vazia — preenchida pelo download)

    Salva em PLANILHA_LISTA_DOWNLOAD e retorna o DataFrame.
    """
    ta_cols_presentes = [c for c in COLUNAS_TA if c in controle_sei.columns]

    controle = controle_sei[
        [COL_SEI_SIAFI, COL_SEI_INSTRUMENTO] + ta_cols_presentes
    ].copy()
    controle = controle.rename(columns={
        COL_SEI_SIAFI:       "_siafi_key",
        COL_SEI_INSTRUMENTO: COL_DOC_AUTORIZATIVO,
    })
    controle["_siafi_key"] = controle["_siafi_key"].astype(str)

    sigcon = sigcon_filtrado[[COL_SIGCON_SIAFI]].copy()
    sigcon[COL_SIGCON_SIAFI] = sigcon[COL_SIGCON_SIAFI].astype(str)

    merged = pd.merge(
        sigcon, controle,
        left_on=COL_SIGCON_SIAFI,
        right_on="_siafi_key",
        how="left",
    ).drop(columns=["_siafi_key"])

    merged = merged.drop_duplicates(subset=[COL_SIGCON_SIAFI]).reset_index(drop=True)
    merged[COL_OBSERVACAO] = ""

    merged.to_excel(PLANILHA_LISTA_DOWNLOAD, index=False)
    print(f"✅ Lista de download gerada — {len(merged)} instrumento(s) → {PLANILHA_LISTA_DOWNLOAD.name}")
    return merged


# ── Carga das bases ───────────────────────────────────────────────────────────

def carregar_bases():
    """Carrega Controle SEI.xlsx e Consultas SIGCON.xlsx."""
    controle_sei     = pd.read_excel(PLANILHA_CONTROLE)
    consultas_sigcon = pd.read_excel(PLANILHA_CONSULTAS_SIGCON)
    print(f"✅ Bases carregadas — Controle SEI: {len(controle_sei)} linhas | SIGCON: {len(consultas_sigcon)} linhas")
    print(f"ℹ️  Colunas SIGCON: {consultas_sigcon.columns.tolist()}")
    return controle_sei, consultas_sigcon


# ── Filtro do SIGCON ──────────────────────────────────────────────────────────

def filtrar_sigcon(df):
    """
    Filtra o SIGCON:
      - Remove situação BLOQUEADO
      - Mantém apenas instrumentos publicados nos últimos 3 anos (inclusive ano atual)
    """
    df = df.copy()
    df[COL_SIGCON_DATA_PUB] = pd.to_datetime(
        df[COL_SIGCON_DATA_PUB], dayfirst=True, errors="coerce"
    )

    df = df[
        df[COL_SIGCON_DATA_PUB].dt.year.isin(DATA_FILTRO) &
        (df[COL_SIGCON_SITUACAO] != SITUACAO_BLOQUEADO)
    ].copy()

    df[COL_SIGCON_DATA_PUB] = df[COL_SIGCON_DATA_PUB].dt.strftime("%d/%m/%Y")

    print(
        f"✅ SIGCON filtrado — {len(df)} registros "
        f"(NOT BLOQUEADO, {DATA_FILTRO[0]}–{DATA_FILTRO[-1]})."
    )
    return df


# ── Cruzamento de dados ───────────────────────────────────────────────────────

def cruzar_dados(sigcon_filtrado, controle_sei, lista_instrumentos, lista_ta_drive):
    """
    Cruza quatro fontes de dados:
      1. SIGCON filtrado  ←[SIAFI]→  Controle SEI  (adiciona Doc_autorizativo + colunas TA)
      2. Resultado        ←[Doc_autorizativo]→  Drive Instrumento  (nome_pdf, id.y, drive_resource)
      3. Colunas TA       → substituídas pelo link do Drive 'Termos aditivos'
    Também constrói a URL do TransfereGov a partir do Código.União.
    """
    # ── Prepara Controle SEI ──────────────────────────────────────────────────
    ta_cols_presentes = [c for c in COLUNAS_TA if c in controle_sei.columns]
    controle = controle_sei[
        [COL_SEI_SIAFI, COL_SEI_INSTRUMENTO] + ta_cols_presentes
    ].copy()
    controle = controle.rename(columns={
        COL_SEI_SIAFI:       "_siafi_key",
        COL_SEI_INSTRUMENTO: COL_DOC_AUTORIZATIVO,
    })
    controle["_siafi_key"] = controle["_siafi_key"].astype(str)

    # ── Prepara SIGCON ────────────────────────────────────────────────────────
    sigcon = sigcon_filtrado.copy()
    sigcon[COL_SIGCON_SIAFI] = sigcon[COL_SIGCON_SIAFI].astype(str)

    # Constrói URL do TransfereGov; deixa em branco quando Código União for 0 ou vazio
    cod_uniao = sigcon[COL_SIGCON_CODIGO_UNIAO]
    tem_codigo = cod_uniao.notna() & (cod_uniao.astype(str).str.strip() != "0") & (cod_uniao.astype(str).str.strip() != "")
    sigcon[COL_SIGCON_INTEIRO_TRANSFERE] = None
    sigcon.loc[tem_codigo, COL_SIGCON_INTEIRO_TRANSFERE] = (
        URL_TRANSFEREGOV_BASE
        + cod_uniao[tem_codigo].astype(str)
        + "&Usr=guest&Pwd=guest"
    )

    # ── Join 1: SIGCON + Controle SEI ────────────────────────────────────────
    merged = pd.merge(
        sigcon, controle,
        left_on=COL_SIGCON_SIAFI,
        right_on="_siafi_key",
        how="left",
    )
    merged = merged.drop(columns=["_siafi_key"], errors="ignore")

    # ── Join 2: resultado + Drive ─────────────────────────────────────────────
    drive_cols = lista_instrumentos[
        [COL_DOC_AUTORIZATIVO, "name", "id", COL_DRIVE_RESOURCE]
    ].copy()

    merged = pd.merge(merged, drive_cols, on=COL_DOC_AUTORIZATIVO, how="left")

    # Renomeia colunas do Drive para os nomes finais esperados
    merged = merged.rename(columns={"name": COL_NOME_PDF, "id": COL_ID_PDF})

    # ── Join 3: substitui valores numéricos das colunas TA por links do Drive ─
    if not lista_ta_drive.empty and ta_cols_presentes:
        link_por_stem = lista_ta_drive.set_index("stem")["drive_resource_ta"]
        for ta_col in ta_cols_presentes:
            sufixo = ta_col.split()[-1]
            def _link_ta(val, _suf=sufixo, _lookup=link_por_stem):
                if pd.isna(val):
                    return None
                stem = f"{int(float(val))}_{_suf}"
                return _lookup.get(stem)
            merged[ta_col] = merged[ta_col].apply(_link_ta)

    print(f"✅ Cruzamento concluído — {len(merged)} registros.")
    return merged


# ── Preenchimento do inteiro teor (SIGCON) ────────────────────────────────────

def preencher_link_inteiro_teor(df, service, pasta_id):
    """
    Preenche inteiro teor do Sigcon com o link do Drive quando estiver vazio.

    Etapa 1 — usa drive_resource já trazido pelo join (rápido, sem chamada extra).
    Etapa 2 — para linhas que ainda ficaram sem link mas têm Doc_autorizativo,
               busca o arquivo no Drive por nome ({Doc_autorizativo}.pdf),
               obtém o id e monta o link direto.
    """
    if COL_SIGCON_INTEIRO_SIGCON not in df.columns:
        df[COL_SIGCON_INTEIRO_SIGCON] = None

    # Etapa 1: join já trouxe drive_resource
    mask1 = df[COL_SIGCON_INTEIRO_SIGCON].isna() & df[COL_DRIVE_RESOURCE].notna()
    df.loc[mask1, COL_SIGCON_INTEIRO_SIGCON] = df.loc[mask1, COL_DRIVE_RESOURCE]
    print(f"✅ Etapa 1 — {mask1.sum()} link(s) inserido(s) via join.")

    # Etapa 2: busca explícita no Drive para linhas ainda sem link
    mask2 = df[COL_SIGCON_INTEIRO_SIGCON].isna() & df[COL_DOC_AUTORIZATIVO].notna()
    pendentes = df.loc[mask2, COL_DOC_AUTORIZATIVO].unique()
    print(f"ℹ️  Etapa 2 — buscando {len(pendentes)} arquivo(s) no Drive por nome...")

    cache_ids = {}
    for doc in pendentes:
        nome_pdf = f"{doc}.pdf"
        try:
            resultado = service.files().list(
                q=f"name='{nome_pdf}' and '{pasta_id}' in parents and trashed=false",
                fields="files(id, name)",
                pageSize=1,
            ).execute()
            arquivos = resultado.get("files", [])
            if arquivos:
                file_id = arquivos[0]["id"]
                link = f"https://drive.google.com/file/d/{file_id}/view"
                cache_ids[doc] = link
                idx = df.index[mask2 & (df[COL_DOC_AUTORIZATIVO] == doc)]
                df.loc[idx, COL_ID_PDF]        = file_id
                df.loc[idx, COL_NOME_PDF]       = nome_pdf
                df.loc[idx, COL_DRIVE_RESOURCE] = link
        except Exception as e:
            print(f"  ⚠️  Erro ao buscar '{nome_pdf}' no Drive: {e}")

    df.loc[mask2, COL_SIGCON_INTEIRO_SIGCON] = (
        df.loc[mask2, COL_DOC_AUTORIZATIVO].map(cache_ids)
    )
    encontrados = df.loc[mask2, COL_SIGCON_INTEIRO_SIGCON].notna().sum()
    print(f"✅ Etapa 2 — {encontrados} link(s) encontrado(s) via busca no Drive.")
    return df


# ── Exportação ────────────────────────────────────────────────────────────────

def exportar_planilha(df):
    """Exporta o DataFrame final para .xlsx com formatação Tahoma 8pt e bordas."""
    # Reordena colunas conforme COLUNAS_FINAIS; colunas ausentes são ignoradas
    colunas_presentes = [c for c in COLUNAS_FINAIS if c in df.columns]
    df = df[colunas_presentes]

    wb = Workbook()
    ws = wb.active
    ws.title = ABA_EXCEL

    borda = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    estilo_cabecalho = Font(name=EXCEL_FONTE, size=EXCEL_TAMANHO, bold=True)
    estilo_dados     = Font(name=EXCEL_FONTE, size=EXCEL_TAMANHO)
    alinha_centro    = Alignment(horizontal="center", vertical="center")
    alinha_esquerda  = Alignment(horizontal="left",   vertical="center")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = estilo_cabecalho
        cell.alignment = alinha_centro
        cell.border    = borda
        ws.column_dimensions[get_column_letter(col_idx)].width = EXCEL_LARGURA_COL

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = estilo_dados
            cell.alignment = alinha_esquerda
            cell.border    = borda

    nome_arquivo = f"{PLANILHA_OUTPUT_PREFIX} {datetime.now().strftime('%d-%m-%Y')}.xlsx"
    caminho      = CAMINHO_FINAL / nome_arquivo
    wb.save(caminho)
    print(f"✅ Planilha exportada: {nome_arquivo}")


# ── Execução direta (sem filtro pré-aplicado) ─────────────────────────────────

if __name__ == "__main__":
    controle_sei, consultas_sigcon = carregar_bases()
    sigcon_filtrado    = filtrar_sigcon(consultas_sigcon)
    gerar_lista_download(sigcon_filtrado, controle_sei)

    service            = autenticar_drive()
    fazer_upload_pdfs(service)
    pasta_id           = _id_pasta_drive(service)
    lista_instrumentos = listar_pdfs_drive(service)
    lista_ta_drive     = listar_pdfs_ta_drive(service)
    df = cruzar_dados(sigcon_filtrado, controle_sei, lista_instrumentos, lista_ta_drive)
    df = preencher_link_inteiro_teor(df, service, pasta_id)
    exportar_planilha(df)
