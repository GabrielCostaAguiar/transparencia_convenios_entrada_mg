import os #biblioteca para acessar o proxy
os.environ['HTTP_PROXY'] = "http://m1578465:Dcgce2025b767Gb@proxycamg.prodemge.gov.br:8080"
os.environ['HTTPS_PROXY'] = "http://m1578465:Dcgce2025b@proxycamg.prodemge.gov.br:8080"
os.environ['NO_PROXY'] = "localhost,127.0.0.1"

from selenium import webdriver # Biclioteca de automações web
from webdriver_manager.chrome import ChromeDriverManager #download do chromedriver atualizado / Problema para funcionar devido ao Proxy
from selenium.webdriver.chrome.service import Service #abrindo biblioteca
from selenium.webdriver.chrome.options import Options
import time #biblioteca para fazer o robo esperar a página carregar
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC #Biblioteca para if de esperar elementos da pagina carregar
from selenium.webdriver.common.by import By #BIBLIOTECA PARA AJUDAR A ACESSAR O IFRAME
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.common.alert import Alert #para lidar com o alerta javascript da visualização da nota técnica
import requests #biblioteca para verificar resposta do servidor
import pandas as pd #ler, baixar e manipular planilhas e base de dados
from selenium.common.exceptions import NoSuchElementException #para verificar anulação
from selenium.common.exceptions import TimeoutException
import base64 #usado para salvar o pdf direto da pagina da web

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter #Bibliotecas para formatar planilha

# Caminhos e pastas
download_dir = r"C:\Users\M1578465\Documents\DCGCE\Automacoes\Transparência SIGCON"
os.makedirs(download_dir, exist_ok=True)  # cria a pasta se não existir

# Diretório onde os PDFs serão salvos
diretorio_base = r"C:\Users\M1578465\Documents\DCGCE\Automacoes\Transparência SIGCON\PDFs"
os.makedirs(diretorio_base, exist_ok=True)

# Carrega a planilha
controle_sei = pd.read_excel(os.path.join(download_dir, "Controle SEI.xlsx"))

# Pega coluna 1 e 5 e cria uma lista de tuplas [(num_instrumento, valor_pesquisa), ...]
instrumentos = controle_sei[['Nº_SEI', 'Instrumento']].dropna(subset=['Instrumento'])
print(instrumentos.info())

# Configuração Selenium
caminho_driver = r"C:\Users\M1578465\.wdm\drivers\chromedriver\win64\141.0.7390.65\chromedriver-win32\chromedriver.exe"
options = Options()
options.add_argument("--start-maximized")
servico = Service(caminho_driver)
navegador = webdriver.Chrome(service=servico, options=options)

# Acessa o SEI
url_sei = "https://www.sei.mg.gov.br/sip/login.php?sigla_orgao_sistema=GOVMG&sigla_sistema=SEI&infra_url=L3NlaS8="
navegador.get(url_sei)

# XPaths
xpath_usuario = '//*[@id="txtUsuario"]'
xpath_senha = '//*[@id="pwdSenha"]'
xpath_botao = '//*[@id="Acessar"]'
xpath_tela_login = '//*[@id="area-cards-login"]'
xpath_orgao = '//*[@id="selOrgao"]'
xpath_scc = '//*[@id="selOrgao"]/option[52]'
# Login
WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, xpath_tela_login)))
usuario = navegador.find_element(By.XPATH, xpath_usuario)
senha = navegador.find_element(By.XPATH, xpath_senha)
botao = navegador.find_element(By.XPATH, xpath_botao)
orgao = navegador.find_element(By. XPATH, xpath_orgao)
orgao_scc = navegador.find_element(By. XPATH, xpath_scc)

usuario.send_keys("14377385674")
senha.send_keys("88073227Gb!")
orgao.click()

# Seleciona órgão
orgao_scc.click()  # 52ª opção

time.sleep(1)
botao.click()

WebDriverWait(navegador, 10).until_not(EC.presence_of_element_located((By.XPATH, xpath_tela_login)))
print("✅ Login efetuado com sucesso!")

#Fechando imagem pós login
# Espera até a imagem de fechar aparecer
try:
    body_pagina = '/html/body'
    fechar_imagem = WebDriverWait(navegador, 10).until(
        EC.element_to_be_clickable((By.XPATH, body_pagina))
    )
    fechar_imagem.send_keys('\ue00c')
    print("✅ Modal fechada com sucesso!")
except Exception as e:
    print(f"❌ Erro ao fechar a modal: {e}")

# Loop pelos instrumentos
xpath_pesquisa = '//*[@id="txtPesquisaRapida"]'
xpath_download = '//*[@id="icon"]/cr-icon' # ajuste de acordo com o botão na página
enter = '\ue007'
# '\ue007' código tecla enter

# Loop pelos instrumentos usando DataFrame
for idx, row in instrumentos.iterrows():
    num_sei = row['Nº_SEI']
    instrumento = row['Instrumento']  # <-- este será usado no pesquisar
    try:
        # Volta ao DOM principal antes de cada pesquisa
        navegador.switch_to.default_content()

        # Limpa o campo e pesquisa pelo instrumento
        campo_pesquisa = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath_pesquisa))
        )
        campo_pesquisa.send_keys(str(instrumento))
        campo_pesquisa.send_keys(enter)
        
        time.sleep(3)
        try:
            #Acessando o iframe da página Nota Técnica / Outras Funções
            iframe_doc = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ifrVisualizacao"]')) ) 
            navegador.switch_to.frame(iframe_doc)
            print("iframe do doc acessado com sucesso")
        except:
            print("Erro ao acessar o iframe doc")
            continue
        
        try:
            # Encontra o link do PDF
            link_element = WebDriverWait(navegador, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="divArvoreInformacao"]/a'))
            )
            href = link_element.get_attribute('href')

            # Monta URL completa (adapte o base_url se necessário)
            base_url = "https://www.sei.mg.gov.br/sei/controlador.php?"
            pdf_url = base_url + href.split("controlador.php?")[1]

            # Pega cookies da sessão Selenium
            cookies = {c['name']: c['value'] for c in navegador.get_cookies()}

            # Faz download do PDF
            caminho_pdf = os.path.join(diretorio_base, f"{instrumento}.pdf")
            # Verifica se o PDF já existe
            if os.path.exists(caminho_pdf):
                print(f"⚠️ PDF do {num_sei} já existe, pulando...")
                continue
            
            response = requests.get(pdf_url, cookies=cookies)
            with open(caminho_pdf, 'wb') as f:
                f.write(response.content)

            print(f"✅ PDF do {num_sei} baixado com sucesso!")
        except Exception as e:
            print(f"❌ Erro ao fazer o download do {instrumento}")
            continue

    except Exception as e:
        print(f"❌ Erro ao processar {num_sei} ({instrumento}): {e}")